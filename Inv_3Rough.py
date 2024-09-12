
import openpyxl
from openpyxl import cell

inv_temp_wb = openpyxl.load_workbook("Source/invoice template.xlsx")

# inv_temp_wb.save('C:\\Users\\panka\\OneDrive\\Desktop\\Abcd.xlsx')

# Iterating column and row wise in invoice template

ws = inv_temp_wb["Invoice"]

for i in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for j in i:
        # print("Sheet",j.value)
        if j.value == "Customer ID":
            # Cell D9 will be the cell in which value of Customer ID will be enter for a customer.
            D9 = ws.cell(row=j.row, column=j.column+1)
            print(D9)            
            # print(j)

for i in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for j in i:
        if j.value == "Customer Name":
            # Cell D10 will be the cell in which value of Customer Name will be enter for a customer.
            D10 = ws.cell(row=j.row, column=j.column+1)
            print(D10)            
            # print(j)

for i in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for j in i:
        if j.value == "Segment":
            # Cell D11 will be the cell in which value of Segment will be enter for a customer.
            D11 = ws.cell(row=j.row, column=j.column+1)
            print(D11)            
            # print(j)

for i in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for j in i:
        if j.value == "Address":
            # Cell D12 will be the cell in Address will be enter for a customer.
            D12 = ws.cell(row=j.row, column=j.column+1)
            print(D12)            
            # print(j)

for i in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for j in i:
        if j.value == "Order Details":
            # Cell D15 will be the cell in which Order Details will be enter for a customer.
            D15 = ws.cell(row=j.row, column=j.column+1)
            print(D15)            
            # print(j)


for i in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for j in i:
        if j.value == "Order ID":
            # Cell M9 will be the cell in which Order ID will be enter for a customer.
            M9 = ws.cell(row=j.row, column=j.column+1)
            print(M9)            
            # print(j)


for i in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for j in i:
        if j.value == "Order Date":
            # Cell M10 will be the cell in which Order Date will be enter for a customer.
            M10 = ws.cell(row=j.row, column=j.column+1)
            print(M10)            
            # print(j)


for i in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for j in i:
        if j.value == "Ship Date":
            # Cell M11 will be the cell in which Ship Date will be enter for a customer.
            M11 = ws.cell(row=j.row, column=j.column+1)
            print(M11)            
            # print(j)


for i in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for j in i:
        if j.value == "Product ID":
            # Cell C18 will be the cell in which Product ID will be enter for a customer.
            C18 = ws.cell(row=j.row+1, column=j.column)
            print(C18)            
            # print(j)

for i in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for j in i:
        if j.value == "Product Name":
            # Cell D18 will be the cell in which Product Name will be enter for a customer.
            D18 = ws.cell(row=j.row+1, column=j.column)
            print(D18)            
            # print(j)

for i in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for j in i:
        if j.value == "Unit Price":
            # Cell E18 will be the cell in which Unit Price will be enter for a customer.
            E18 = ws.cell(row=j.row+1, column=j.column)
            print(E18)            
            # print(j)

for i in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for j in i:
        if j.value == "Quantity":
            # Cell F18 will be the cell in which Quantity will be enter for a customer.
            F18 = ws.cell(row=j.row+1, column=j.column)
            print(F18)            
            # print(j)

for i in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for j in i:
        if j.value == "Price":
            # Cell F18 will be the cell in which Price will be enter for a customer.
            F18 = ws.cell(row=j.row+1, column=j.column)
            print(F18)            
            # print(j)


# I will be converting this into a function in Inv_3Rough.py




