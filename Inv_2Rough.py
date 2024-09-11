
import openpyxl
inv_temp_wb = openpyxl.load_workbook("Source/invoice template.xlsx")

# inv_temp_wb.save('C:\\Users\\panka\\OneDrive\\Desktop\\Abcd.xlsx')

# Iterating column and row wise in invoice template

for i in inv_temp_wb["Invoice"].iter_rows(min_row=1, max_row=inv_temp_wb["Invoice"].max_row, min_col=1, max_col=inv_temp_wb["Invoice"].max_column):
    for j in i:
        print("Sheet",j.value)
        if j.value == "Customer ID":
            print

