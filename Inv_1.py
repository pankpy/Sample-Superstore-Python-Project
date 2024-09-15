import pandas as pd
import openpyxl
# from openpyxl.drawing.image import Image
import pillow

import warnings


# warnings.filterwarnings("ignore")


# Reading sample superstore dataset
data_file_path = 'E:\\zPankaj\\Sample Superstore Invoices project\\Source\\Sample - Superstore.xlsx'
df = pd.read_excel(data_file_path, sheet_name='Orders')

print("Hello1",df)

df.drop_duplicates(inplace=True)


#Keeping only these columns
df = df[['Order ID', 'Order Date', 'Ship Date', 'Region',
       'Customer ID', 'Customer Name', 'Segment', 'Country', 'City', 'State',
       'Postal Code','Product ID',
       'Product Name', 'Sales', 'Quantity', 'Discount', 'Profit']]

# Grouping Data 
# Order ID wise


#Accessing invoice template and pasting data from df1 to invoice template.
inv_temp_wb = openpyxl.load_workbook("Source/invoice template.xlsx")
ws = inv_temp_wb["Invoice"]

# logo_path = 'E:\\zPankaj\\Sample Superstore Invoices project\\Source'
# logo_image = Image(logo_path)  # Replace with the path to the image
# ws.add_image(logo_image, 'A1')

# inv_temp_wb.save('C:\\Users\\panka\\OneDrive\\Desktop\\Abcd.xlsx')

# Iterating column and row wise in invoice template
# function to find cell next to Or below required fields in invoice.
def cell_for_entering_value_finder(inv_field):
    for i in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for j in i:
            # print("Sheet",j.value)
            if j.value == inv_field:
                if inv_field in ["Order Details"]:
                # Cell_num will be the cell in which value of inv field will be enter for a customer.
                    cell_num = ws.cell(row=j.row+1, column=j.column)
                else:
                    cell_num = ws.cell(row=j.row, column=j.column+1)

    return cell_num


#Creating list of distinct  Order ID
order_id_list = df['Order ID'].unique().tolist()
# print(len(order_id_list))

#Note: There are 5009 distinct Order ID. It would create 5009 invoices in Invoices folder and would take a lot of space. 
#For demonstration, I am adding a Region filter
df = df[df['Region'] == 'West']

def invoice_details(df1):

    try:
        # print(df1)
        cust_name = (df1["Customer Name"].unique().tolist())[0]
        cust_ID = (df1["Customer ID"].unique().tolist())[0]
        segment = (df1["Segment"].unique().tolist())[0]
        order_ID = (df1["Order ID"].unique().tolist())[0]
        df1["Order Date"] = pd.to_datetime(df1["Order Date"], errors='coerce')
        df1["Order Date"] = df1["Order Date"].apply(lambda x: x.strftime("%d %b %Y"))

        order_dt = (df1["Order Date"].unique().tolist())[0]

        df1["Ship Date"] = pd.to_datetime(df1["Ship Date"], errors='coerce')
        df1["Ship Date"] = df1["Ship Date"].apply(lambda x: x.strftime("%d %b %Y"))

        ship_dt = (df1["Ship Date"].unique().tolist())[0]
        
        prod_ID = (df1["Product ID"].unique().tolist())[0]
        prod_name = (df1["Product Name"].unique().tolist())[0]
        qty = (df1["Quantity"].unique().tolist())[0]

        #Address - concat three columns
        df1['Address'] = df1['City'] + ', '+ df1['State'] +', '+ df1['Country']
        address = (df1["Address"].unique().tolist())[0]
        # print(cust_name,cust_ID, segment, order_ID,order_dt,ship_dt,prod_ID,prod_name,qty)

    except Exception as e:
        print("Error: ",e)

    return {"Customer Name": cust_name,"Customer ID":cust_ID, "Segment":segment, "Order ID":order_ID,
            "Order Date":order_dt,"Ship Date":ship_dt,"Product ID":prod_ID,"Product Name":prod_name,
            "Quantity":qty, "Address":address}

    # return cust_name,cust_ID, segment, order_ID,order_dt,ship_dt,prod_ID,prod_name,qty


def fill_invoice(df1,**kwargs):
   
    for i_field in kwargs:

        if i_field not in ["Product ID", "Product Name", "Quantity"]:

            cell_inv = cell_for_entering_value_finder(i_field)
            cell_inv.value = kwargs[i_field]

    # Saving this in New invoice file...
    cn = kwargs["Customer Name"]
    cstID = kwargs["Customer ID"]
    invoice_name = f"{cn}-{cstID}.xlsx"
    new_inv_path = f'E:\\zPankaj\\Sample Superstore Invoices project\\Invoices\\{invoice_name}'
    inv_temp_wb.save(new_inv_path)
 
    # Order details
    # Keeping these fields in invoice
    col = ["Product ID", "Product Name", "Quantity", "Sales", "Discount"]

    # Where I want to paste this in invoice template
    invf = "Order Details"
    cell_inv = cell_for_entering_value_finder(invf)

    df2 = df1[col]

    # Adding one more column "Unit Price". Value for this will be calculated by subtracting Discount from Sales, 
    # then dividing the difference by Quantity.

    # Compute the Unit Price
    df2['Unit Price'] = (df2["Sales"] - df2["Discount"]) / df2['Quantity']
    df2['Unit Price'] = round(df2['Unit Price'],2)
    df2['Sales'] = round(df2['Sales'],2)
    df2 = df2[["Product ID", "Product Name", "Unit Price", "Quantity", "Discount", "Sales"]]

    # Adding Total under Sales column
    total_sales = df2['Sales'].sum()
    total_row_df = pd.DataFrame({'Product ID': ['Total'], 'Product Name': [''], 'Quantity': [''], 'Sales': [total_sales]})
    df2 = pd.concat([df2, total_row_df], ignore_index=True)


    # Currently, these two sentences are already in invoice template. The can be added in code to add after Total row dynamically.
    # Thank you for shopping with us.
    # This is a system generated receipt, hence signature not required.


    start_cell = cell_inv
    # breakpoint()
    # start_row = ws[start_cell].row
    start_row = start_cell.row+1
    start_column = start_cell.column

    # saving in new invoice...
    wb2 = openpyxl.load_workbook(new_inv_path)
    ws2 = wb2["Invoice"]
    # Write the DataFrame to invoice using openpyxl
    for r_idx, row in enumerate(df2.itertuples(index=False), start=start_row):
        for c_idx, value in enumerate(row, start=start_column):
            ws2.cell(row=r_idx, column=c_idx, value=value)
    wb2.save(new_inv_path)


def convert_to_pdf():
    # The newly created invoice can be converted into pdf.
    # Use the new invoice path
    pass
 
for ord in order_id_list:
    df1 = df[df['Order ID'] == ord]
     # print('DF1', df1)
    if (not df1.empty) or (not df1.isna().all().all()):
        # fill_invoice(df1)

        kwargs = invoice_details(df1)
        # print("Here is the output: ", kwargs)
        try:
            pass
            fill_invoice(df1,**kwargs)
        except Exception as e:
            print("Capture Error in log., Error is: ",e)